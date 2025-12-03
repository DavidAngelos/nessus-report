#!/usr/bin/env python3
"""
Nessus CSV Report Generator
Converts Nessus CSV exports into professional customer-ready reports
"""

import pandas as pd
import argparse
import sys
from datetime import datetime
from pathlib import Path
import re
from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

class NessusReportGenerator:
    def __init__(self, csv_file):
        self.csv_file = csv_file
        self.df = None
        self.summary_stats = {}
        
    def load_data(self):
        """Load and validate the Nessus CSV file"""
        try:
            # Try different encodings in case of special characters
            encodings = ['utf-8', 'latin-1', 'cp1252']
            for encoding in encodings:
                try:
                    self.df = pd.read_csv(self.csv_file, encoding=encoding)
                    print(f"Successfully loaded {len(self.df)} records using {encoding} encoding")
                    break
                except UnicodeDecodeError:
                    continue
            
            if self.df is None:
                raise ValueError("Could not read CSV file with any supported encoding")
                
            # Display available columns for debugging
            print(f"Available columns: {list(self.df.columns)}")
            
            # Clean column names (remove extra spaces, standardize)
            self.df.columns = self.df.columns.str.strip()
            
            # Remove informational findings that are not security issues
            informational_filters = [
                'Nessus Scan Information',
                'Traceroute Information',
                'Service Detection',
                'Nessus SYN scanner',
                'OS Identification',
                'Device Type',
                'Common Platform Enumeration',
                'Target Credential Status',
                'OS Security Patch Assessment Not Available'
            ]
            
            print(f"Total findings before filtering: {len(self.df)}")
            
            # Filter out purely informational findings
            self.df = self.df[~self.df['Name'].isin(informational_filters)]
            
            print(f"Security-relevant findings after filtering: {len(self.df)}")
            
            return True
            
        except Exception as e:
            print(f"Error loading CSV file: {e}")
            return False
    
    def clean_data(self):
        """Clean and standardize the data"""
        # Remove empty rows
        self.df = self.df.dropna(how='all')
        
        # Clean up text fields - remove extra quotes and whitespace
        text_columns = ['Name', 'Synopsis', 'Description', 'Solution', 'CVE']
        for col in text_columns:
            if col in self.df.columns:
                self.df[col] = (
                    self.df[col]
                    .astype(str)
                    .str.strip()
                    .str.replace('"', '')
                )
        
        # Clean up Risk field
        if 'Risk' in self.df.columns:
            self.df['Risk'] = self.df['Risk'].astype(str).str.strip()

            # ✅ Keep only Low, Medium, High, Critical
            valid_risks = ['Low', 'Medium', 'High', 'Critical']
            before_count = len(self.df)
            self.df = self.df[self.df['Risk'].isin(valid_risks)]
            print(f"Filtered by severity: {before_count} → {len(self.df)} relevant issues")
        
        # Clean up CVSS scores
        for col in ['CVSS v2.0 Base Score', 'CVSS v3.0 Base Score']:
            if col in self.df.columns:
                self.df[col] = self.df[col].astype(str).str.strip()
                # Convert empty strings to None
                self.df[col] = self.df[col].replace('', None)
        
        # Clean up Host/IP information
        if 'Host' in self.df.columns:
            self.df['Host'] = self.df['Host'].astype(str).str.strip()
            
        # Create a combined CVSS score column for sorting
        self.df['CVSS_Score'] = None
        if 'CVSS v3.0 Base Score' in self.df.columns:
            self.df['CVSS_Score'] = pd.to_numeric(self.df['CVSS v3.0 Base Score'],
                                                errors='coerce')
        
        # Fall back to CVSS v2.0 if v3.0 is not available
        if 'CVSS v2.0 Base Score' in self.df.columns:
            mask = self.df['CVSS_Score'].isna()
            self.df.loc[mask, 'CVSS_Score'] = pd.to_numeric(
                self.df.loc[mask, 'CVSS v2.0 Base Score'],
                errors='coerce'
            )
    
    def generate_summary(self):
        """Generate summary statistics"""
        # Risk level summary
        if 'Risk' in self.df.columns:
            risk_counts = self.df['Risk'].value_counts()
            self.summary_stats['by_severity'] = risk_counts.to_dict()
            
            # Calculate risk score based on severity
            risk_weights = {'Critical': 4, 'High': 3, 'Medium': 2, 'Low': 1, 'None': 0}
            total_risk_score = sum(risk_weights.get(risk, 0) * count for risk, count in risk_counts.items())
            self.summary_stats['risk_score'] = total_risk_score
        
        # Host statistics
        if 'Host' in self.df.columns:
            hosts_with_issues = self.df[self.df['Risk'] != 'None']['Host'].nunique()
            total_hosts = self.df['Host'].nunique()
            self.summary_stats['total_hosts'] = total_hosts
            self.summary_stats['hosts_with_issues'] = hosts_with_issues
            
            # Most affected hosts
            host_counts = self.df[self.df['Risk'] != 'None']['Host'].value_counts()
            self.summary_stats['most_affected_hosts'] = host_counts.head(10).to_dict()
        
        # Total findings
        security_findings = len(self.df[self.df['Risk'] != 'None'])
        self.summary_stats['total_findings'] = len(self.df)
        self.summary_stats['security_findings'] = security_findings
        
        # Top vulnerabilities by occurrence
        vuln_counts = self.df[self.df['Risk'] != 'None']['Name'].value_counts()
        self.summary_stats['top_vulnerabilities'] = vuln_counts.head(10).to_dict()
        
        # CVSS statistics
        if 'CVSS_Score' in self.df.columns:
            cvss_data = self.df[self.df['CVSS_Score'].notna()]['CVSS_Score']
            if len(cvss_data) > 0:
                self.summary_stats['avg_cvss'] = round(cvss_data.mean(), 2)
                self.summary_stats['max_cvss'] = round(cvss_data.max(), 2)
                self.summary_stats['high_cvss_count'] = len(cvss_data[cvss_data >= 7.0])
        
        # Port/Service statistics
        if 'Port' in self.df.columns:
            common_ports = self.df[self.df['Risk'] != 'None']['Port'].value_counts()
            self.summary_stats['common_vulnerable_ports'] = common_ports.head(10).to_dict()
    
    def create_executive_summary(self):
        """Create executive summary dataframe"""
        summary_data = []
        
        # Overview
        summary_data.append(['Assessment Overview', ''])
        summary_data.append(['Total Hosts Scanned', self.summary_stats.get('total_hosts', 'N/A')])
        summary_data.append(['Hosts with Security Issues', self.summary_stats.get('hosts_with_issues', 'N/A')])
        summary_data.append(['Total Findings', self.summary_stats.get('total_findings', 'N/A')])
        summary_data.append(['Security-Relevant Findings', self.summary_stats.get('security_findings', 'N/A')])
        
        # CVSS Statistics
        if 'avg_cvss' in self.summary_stats:
            summary_data.append(['', ''])
            summary_data.append(['CVSS Metrics', ''])
            summary_data.append(['Average CVSS Score', self.summary_stats.get('avg_cvss', 'N/A')])
            summary_data.append(['Highest CVSS Score', self.summary_stats.get('max_cvss', 'N/A')])
            summary_data.append(['High Risk Findings (CVSS ≥ 7.0)', self.summary_stats.get('high_cvss_count', 'N/A')])
        
        # Risk Distribution
        if 'by_severity' in self.summary_stats:
            summary_data.append(['', ''])
            summary_data.append(['Risk Level Distribution', ''])
            risk_order = ['Critical', 'High', 'Medium', 'Low', 'None']
            for risk in risk_order:
                count = self.summary_stats['by_severity'].get(risk, 0)
                if count > 0:
                    summary_data.append([f'{risk} Risk', count])
        
        # Top Issues
        if 'top_vulnerabilities' in self.summary_stats:
            summary_data.append(['', ''])
            summary_data.append(['Most Common Vulnerabilities', ''])
            for vuln, count in list(self.summary_stats['top_vulnerabilities'].items())[:5]:
                # Truncate long vulnerability names
                vuln_name = vuln[:60] + "..." if len(vuln) > 60 else vuln
                summary_data.append([vuln_name, count])
        
        return pd.DataFrame(summary_data, columns=['Metric', 'Value'])
    
    def generate_colored_findings_table(self):
        detailed_df = self.create_detailed_report()

        def color_row(row):
            # Color mapping (your requested colors)
            color_map = {
                "Critical": "A43C3C",
                "High": "D97A5B",
                "Medium": "E5B66C",
                "Low": "9FBF8F",
            }
            bg = color_map.get(row['Risk'], "")
            return f'background-color: {bg}; color: white;' if bg else ''

        styled_df = detailed_df.style.apply(
            lambda row: [color_row(row)] * len(row), axis=1
        )
        return styled_df.to_html()


    def create_detailed_report(self):
        """Create detailed findings report"""
        # Select columns that are relevant for customer reports
        customer_columns = [
            'Host',
            'Port',
            'Protocol', 
            'Name',
            'Risk',
            'CVSS v3.0 Base Score',
            'CVSS v2.0 Base Score',
            'CVE',
            'Synopsis',
            'Description',
            'Solution'
        ]
        
        # Only include columns that exist in the dataframe
        available_columns = [col for col in customer_columns if col in self.df.columns]
        
        # Filter out informational findings for the customer report
        security_df = self.df[self.df['Risk'] != 'None'].copy()
        
        if len(security_df) == 0:
            # If no security findings, show a subset of all findings
            security_df = self.df.copy()
        
        detailed_df = security_df[available_columns].copy()
        
        if 'CVSS_Score' in security_df.columns:
            detailed_df['CVSS_Score'] = security_df['CVSS_Score']

        # Include CVSS_Score temporarily for sorting
        if 'CVSS_Score' in self.df.columns:
            security_df['CVSS_Score'] = self.df.loc[security_df.index, 'CVSS_Score']

        # Sort by risk level and CVSS score
        risk_order = ['Critical', 'High', 'Medium', 'Low', 'None']
        if 'Risk' in detailed_df.columns:
            detailed_df['Risk'] = pd.Categorical(detailed_df['Risk'], categories=risk_order, ordered=True)
            detailed_df = detailed_df.sort_values(['Risk', 'CVSS_Score'], ascending=[True, False], na_position='last')
        
        # Clean up the description and solution fields for better presentation
        if 'Description' in detailed_df.columns:
            detailed_df['Description'] = detailed_df['Description'].str.replace('\n\n', ' ').str.replace('\n', ' ')
            # Limit description length for readability
            detailed_df['Description'] = detailed_df['Description'].apply(
                lambda x: x[:500] + "..." if len(str(x)) > 500 else x
            )
        
        if 'Solution' in detailed_df.columns:
            detailed_df['Solution'] = detailed_df['Solution'].str.replace('\n\n', ' ').str.replace('\n', ' ')
        
        # Create a combined CVSS column for the report
        if 'CVSS v3.0 Base Score' in detailed_df.columns and 'CVSS v2.0 Base Score' in detailed_df.columns:
            detailed_df['CVSS Score'] = detailed_df['CVSS v3.0 Base Score'].fillna(detailed_df['CVSS v2.0 Base Score'])
            # Remove the individual CVSS columns to avoid confusion
            detailed_df = detailed_df.drop(['CVSS v3.0 Base Score', 'CVSS v2.0 Base Score'], axis=1)
        
        # Remove the temporary CVSS_Score column if it exists
        if 'CVSS_Score' in detailed_df.columns:
            detailed_df = detailed_df.drop('CVSS_Score', axis=1)

        return detailed_df
        
    def create_host_summary(self):
        """Create a summary by host"""
        if 'Host' not in self.df.columns:
            return pd.DataFrame()
        
        host_summary = []
        security_df = self.df[self.df['Risk'] != 'None']
        
        for host in sorted(security_df['Host'].unique()):
            host_data = security_df[security_df['Host'] == host]
            
            # Count findings by risk level
            risk_counts = host_data['Risk'].value_counts()
            
            # Calculate highest CVSS score
            cvss_scores = host_data['CVSS_Score'].dropna()
            max_cvss = cvss_scores.max() if len(cvss_scores) > 0 else 'N/A'
            
            # Get most common ports
            common_ports = host_data['Port'].value_counts().head(3)
            ports_str = ', '.join([f"{port}({count})" for port, count in common_ports.items()])
            
            host_summary.append({
                'Host': host,
                'Total Issues': len(host_data),
                'Critical': risk_counts.get('Critical', 0),
                'High': risk_counts.get('High', 0),
                'Medium': risk_counts.get('Medium', 0),
                'Low': risk_counts.get('Low', 0),
                'Highest CVSS': max_cvss,
                'Common Ports': ports_str
            })
        
        return pd.DataFrame(host_summary)
    
    def export_to_csv(self, output_prefix):
        """Export reports to CSV files"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Executive Summary
        summary_df = self.create_executive_summary()
        summary_file = f"{output_prefix}_executive_summary_{timestamp}.csv"
        summary_df.to_csv(summary_file, index=False)
        print(f"Executive summary exported to: {summary_file}")
        
        # Detailed Report
        detailed_df = self.create_detailed_report()
        detailed_file = f"{output_prefix}_detailed_findings_{timestamp}.csv"
        detailed_df.to_csv(detailed_file, index=False)
        print(f"Detailed findings exported to: {detailed_file}")
        
        return summary_file, detailed_file
    
    def export_to_excel(self, output_prefix):
        """Export reports to Excel with multiple sheets, tables, and charts"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = f"{output_prefix}_security_assessment_{timestamp}.xlsx"

        # Step 1: Create Excel file with pandas
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Executive Summary (we'll overwrite layout later)
            summary_df = self.create_executive_summary()
            summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)

            # Host Summary
            host_summary_df = self.create_host_summary()
            if not host_summary_df.empty:
                host_summary_df.to_excel(writer, sheet_name='Host Summary', index=False)

            # Detailed Findings
            detailed_df = self.create_detailed_report()
            detailed_df.to_excel(writer, sheet_name='Security Findings', index=False)

            # Informational (if any)
            info_findings = pd.DataFrame()
            if 'Risk' in self.df.columns:
                info_findings = self.df[self.df['Risk'] == 'None']

            if not info_findings.empty:
                info_columns = ['Host', 'Port', 'Protocol', 'Name', 'Synopsis']
                available_info_cols = [col for col in info_columns if col in info_findings.columns]
                info_findings[available_info_cols].to_excel(
                    writer, sheet_name='Informational', index=False
                )

        # Step 2: Reopen with openpyxl
        wb = load_workbook(excel_file)

        # ---------- Helper: add an Excel Table to a whole sheet ----------
        def add_table(ws, table_name):
            # Need at least header + one data row
            if ws.max_row < 2 or ws.max_column < 1:
                return

            # Fix any accidental "Column1" header
            for cell in ws[1]:
                if cell.value == "Column1":
                    cell.value = " "

            last_col_letter = get_column_letter(ws.max_column)
            ref = f"A1:{last_col_letter}{ws.max_row}"

            table = Table(displayName=table_name, ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium1",   # white theme
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

            # Make header row text white (since header band is dark)
            for cell in ws[1]:
                cell.font = Font(color="FFFFFFFF", bold=True)

        # ---------- Rebuild Executive Summary as 4 simple blocks (no tables) ----------
        if 'Executive Summary' in wb.sheetnames:
            ws = wb['Executive Summary']
            ws.delete_rows(1, ws.max_row)  # clear existing

            stats = self.summary_stats

            # 1. Assessment Overview
            overview_rows = [
                ("Total Hosts Scanned", stats.get('total_hosts', 'N/A')),
                ("Hosts with Security Issues", stats.get('hosts_with_issues', 'N/A')),
                ("Total Findings", stats.get('total_findings', 'N/A')),
                ("Security-Relevant Findings", stats.get('security_findings', 'N/A')),
            ]

            # 2. CVSS Metrics
            cvss_rows = [
                ("Average CVSS Score", stats.get('avg_cvss', 'N/A')),
                ("Highest CVSS Score", stats.get('max_cvss', 'N/A')),
                ("High Risk Findings (CVSS ≥ 7.0)", stats.get('high_cvss_count', 'N/A')),
            ]

            # 3. Risk Level Distribution
            risk_rows = []
            by_sev = stats.get('by_severity', {})
            for label in ['Critical', 'High', 'Medium', 'Low']:
                count = by_sev.get(label, 0)
                if count:
                    risk_rows.append((f"{label} Risk", count))

            # 4. Most Common Vulnerabilities (top 5)
            vuln_rows = []
            top_vulns = stats.get('top_vulnerabilities', {})
            for vuln, count in list(top_vulns.items())[:5]:
                name = vuln[:80] + "..." if len(vuln) > 80 else vuln
                vuln_rows.append((name, count))

            def write_block(start_row, title, rows):
                """Write a simple 2-column block starting at start_row. Returns next free row."""
                if not rows:
                    return start_row
                # Title row
                ws.cell(row=start_row, column=1, value=title)
                ws.cell(row=start_row, column=1).font = Font(bold=True)
                start_row += 1
                # Data rows
                for label, value in rows:
                    ws.cell(row=start_row, column=1, value=label)
                    ws.cell(row=start_row, column=2, value=value)
                    start_row += 1
                # Blank row after block
                return start_row + 1

            row = 1
            row = write_block(row, "Assessment Overview", overview_rows)
            row = write_block(row, "CVSS Metrics", cvss_rows)
            row = write_block(row, "Risk Level Distribution", risk_rows)
            row = write_block(row, "Most Common Vulnerabilities", vuln_rows)

            # Autosize first two columns
            for col in range(1, 3):
                max_len = 0
                col_letter = get_column_letter(col)
                for cell in ws[col_letter]:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_len + 2

        # ---------- Add tables only to data sheets ----------
        if 'Host Summary' in wb.sheetnames:
            add_table(wb['Host Summary'], "TblHostSummary")

        if 'Security Findings' in wb.sheetnames:
            add_table(wb['Security Findings'], "TblSecurityFindings")

        if 'Informational' in wb.sheetnames:
            add_table(wb['Informational'], "TblInformational")

        # ---------- Charts sheet ----------
        chart_ws = wb.create_sheet("Charts")

        # 1. Pie chart – Findings by Risk (muted colors)
        by_severity = self.summary_stats.get("by_severity", {})
        chart_ws.append(["Risk", "Count"])

        severity_order = ["Critical", "High", "Medium", "Low"]
        severities_present = []
        for sev in severity_order:
            count = by_severity.get(sev, 0)
            if count:
                chart_ws.append([sev, count])
                severities_present.append(sev)

        if severities_present:
            last_row = chart_ws.max_row
            pie = PieChart()
            pie.title = "Findings by Risk"

            data = Reference(chart_ws, min_col=2, min_row=1, max_row=last_row)
            labels = Reference(chart_ws, min_col=1, min_row=2, max_row=last_row)

            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showVal = True

            # Muted severity colors
            color_map = {
                "Critical": "A43C3C",  # darker red
                "High": "D97A5B",      # soft red/orange
                "Medium": "E5B66C",    # amber
                "Low": "9FBF8F",       # soft green
            }

            series = pie.series[0]
            series.data_points = []
            for idx, sev in enumerate(severities_present):
                dp = DataPoint(idx=idx)
                dp.graphicalProperties.solidFill = color_map.get(sev, "808080")
                series.data_points.append(dp)

            chart_ws.add_chart(pie, "E2")

        # 2. Bar chart – Top affected hosts (max 5)
        most_hosts_dict = self.summary_stats.get("most_affected_hosts", {})
        host_items = list(most_hosts_dict.items())[:5]

        if host_items:
            chart_ws.append([])
            chart_ws.append(["Host", "Findings"])
            header_row = chart_ws.max_row

            for host, count in host_items:
                chart_ws.append([host, count])

            last_row = chart_ws.max_row

            bar1 = BarChart()
            bar1.title = "Top Affected Hosts"
            bar1.y_axis.title = "Findings"

            data = Reference(chart_ws, min_col=2, min_row=header_row, max_row=last_row)
            labels = Reference(chart_ws, min_col=1, min_row=header_row + 1, max_row=last_row)

            bar1.add_data(data, titles_from_data=True)
            bar1.set_categories(labels)
            chart_ws.add_chart(bar1, f"E{header_row}")

        # 3. Bar chart – Top vulnerabilities (max 5)
        top_vulns_dict = self.summary_stats.get("top_vulnerabilities", {})
        vuln_items = list(top_vulns_dict.items())[:5]

        if vuln_items:
            chart_ws.append([])
            chart_ws.append(["Vulnerability", "Count"])
            header_row = chart_ws.max_row

            for vuln, count in vuln_items:
                label = vuln[:30] + "..." if len(vuln) > 30 else vuln
                chart_ws.append([label, count])

            last_row = chart_ws.max_row

            bar2 = BarChart()
            bar2.title = "Top Vulnerabilities"
            bar2.y_axis.title = "Occurrences"

            data = Reference(chart_ws, min_col=2, min_row=header_row, max_row=last_row)
            labels = Reference(chart_ws, min_col=1, min_row=header_row + 1, max_row=last_row)

            bar2.add_data(data, titles_from_data=True)
            bar2.set_categories(labels)
            chart_ws.add_chart(bar2, f"E{header_row}")

        wb.save(excel_file)
        print(f"Excel report exported to: {excel_file}")
        return excel_file
    
    def export_to_html(self, output_prefix):
        """Export to HTML report"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        html_file = f"{output_prefix}_security_report_{timestamp}.html"
        
        # Create HTML content
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Security Assessment Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; line-height: 1.6; }}
                .header {{ background: linear-gradient(135deg, #2c3e50, #3498db); color: white; padding: 30px; text-align: center; border-radius: 10px; }}
                .summary {{ background-color: #ecf0f1; padding: 20px; margin: 20px 0; border-radius: 8px; }}
                .risk-high {{ background-color: #e74c3c; color: white; padding: 5px; border-radius: 3px; }}
                .risk-medium {{ background-color: #f39c12; color: white; padding: 5px; border-radius: 3px; }}
                .risk-low {{ background-color: #27ae60; color: white; padding: 5px; border-radius: 3px; }}
                .risk-critical {{ background-color: #8e44ad; color: white; padding: 5px; border-radius: 3px; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                th {{ background-color: #3498db; color: white; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                .section {{ margin: 30px 0; }}
                h2 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
                .metric-value {{ font-weight: bold; color: #2980b9; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Security Assessment Report</h1>
                <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            
            <div class="section">
                <h2>Executive Summary</h2>
                <div class="summary">
                    {self.create_executive_summary().to_html(index=False, escape=False, classes='summary-table')}
                </div>
            </div>
            
            <div class="section">
                <h2>Host Summary</h2>
                {self.create_host_summary().to_html(index=False, escape=False) if not self.create_host_summary().empty else '<p>No host-specific data available</p>'}
            </div>
            
            <div class="section">
                <h2>Security Findings</h2>
                {self.generate_colored_findings_table()}
            </div>
            
            <div class="section">
                <h2>Report Notes</h2>
                <p><strong>Risk Levels:</strong></p>
                <ul>
                    <li><span class="risk-critical">Critical</span>: Immediate action required</li>
                    <li><span class="risk-high">High</span>: Address as soon as possible</li>
                    <li><span class="risk-medium">Medium</span>: Address in next maintenance window</li>
                    <li><span class="risk-low">Low</span>: Address when convenient</li>
                </ul>
                <p><strong>CVSS Scoring:</strong> Common Vulnerability Scoring System scores range from 0.0 to 10.0, with higher scores indicating more severe vulnerabilities.</p>
            </div>
        </body>
        </html>
        """
        
        with open(html_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"HTML report exported to: {html_file}")
        return html_file

def main():
    parser = argparse.ArgumentParser(description='Convert Nessus CSV exports to customer-ready reports')
    parser.add_argument('csv_file', help='Path to the Nessus CSV file')
    parser.add_argument('-o', '--output', default='security_report', help='Output file prefix (default: security_report)')
    parser.add_argument('-f', '--format', choices=['csv', 'excel', 'html', 'all'], default='all', 
                        help='Output format (default: all)')
    
    args = parser.parse_args()
    
    # Check if input file exists
    if not Path(args.csv_file).exists():
        print(f"Error: Input file '{args.csv_file}' not found")
        sys.exit(1)
    
    # Create report generator
    generator = NessusReportGenerator(args.csv_file)
    
    # Load and process data
    if not generator.load_data():
        sys.exit(1)
    
    generator.clean_data()
    generator.generate_summary()
    
    # Export based on format choice
    if args.format in ['csv', 'all']:
        generator.export_to_csv(args.output)
    
    if args.format in ['excel', 'all']:
        generator.export_to_excel(args.output)
    
    if args.format in ['html', 'all']:
        generator.export_to_html(args.output)
    
    print("\nReport generation completed successfully!")

if __name__ == "__main__":
    main()